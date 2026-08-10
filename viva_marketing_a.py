# -*- coding: utf-8 -*-
"""
GLIM Project Viva Examiner
==========================
Share this file with faculty along with the SETUP GUIDE below.

━━━ FACULTY SETUP GUIDE ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

STEP 1 — Install dependencies
    pip install streamlit groq pymupdf python-docx openpyxl requests

STEP 2 — Create your secrets file
    Create a folder called  .streamlit  next to this script.
    Inside it create a file called  secrets.toml  with this content:

        GROQ_API_KEY        = "your-groq-api-key-here"
        GSHEETS_WEBHOOK_URL = "your-google-apps-script-webhook-url-here"

    How to get GROQ_API_KEY    : console.groq.com → API Keys → Create
    How to get GSHEETS_WEBHOOK : See the Apps Script setup guide shared
                                  separately (deploy as web app → copy URL)

STEP 3 — Configure the viva below (lines marked CONFIGURE)
    Set subject, difficulty, number of questions, and time limit.

STEP 4 — Run
    streamlit run viva_project_groq.py

STEP 5 — Share the URL with students
    Use Cloudflare Tunnel (no account needed):
        cloudflared tunnel --url http://localhost:8501
    Or ngrok:
        ngrok http 8501

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

# ==============================================================================
# ⚙️  CONFIGURE BEFORE DEPLOYING  (edit these 4 lines only)
# ==============================================================================
VIVA_SUBJECT           = "Marketing Management"   # Subject name
VIVA_DIFFICULTY        = "Standard"               # Easy | Standard | Rigorous
VIVA_TOTAL_QUESTIONS   = 10                       # Questions before auto-close
VIVA_TIME_LIMIT_MINUTES = 10                      # Session time limit in minutes
# ==============================================================================

import streamlit as st
import streamlit.components.v1 as components
from groq import Groq
import fitz
from docx import Document
import io
import re
import time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Save location ─────────────────────────────────────────────────────────────
# Scores are saved to a local Excel file in the same folder as this script
# AND posted to Google Sheets via the webhook URL in secrets.toml.
try:
    _SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    _SCRIPT_DIR = os.getcwd()
SCORESHEET_FILE = os.path.join(_SCRIPT_DIR, "GLIM_Viva_Scores.xlsx")

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="GLIM Project Viva Examiner", page_icon="🎓")
st.title("🎓 GLIM Project Viva Examiner")
st.caption("Powered by Groq + Llama 3.1 · Oral Viva Mode")

# ── Secrets — loaded from .streamlit/secrets.toml ────────────────────────────
# Faculty: make sure both keys are present in your secrets.toml file.
# The app shows a clear error message if either key is missing.
def _get_secret(key):
    val = st.secrets.get(key) or os.environ.get(key, "")
    return val.strip()

GROQ_API_KEY        = _get_secret("GROQ_API_KEY")
GSHEETS_WEBHOOK_URL = _get_secret("GSHEETS_WEBHOOK_URL")

if not GROQ_API_KEY:
    st.error(
        "⚠️ **GROQ_API_KEY not found.**\n\n"
        "Create `.streamlit/secrets.toml` next to this script and add:\n\n"
        "```toml\n"
        'GROQ_API_KEY = "your-groq-key-here"\n'
        "```\n\n"
        "Get a free key at [console.groq.com](https://console.groq.com)"
    )
    st.stop()

if not GSHEETS_WEBHOOK_URL:
    st.warning(
        "⚠️ **GSHEETS_WEBHOOK_URL not found in secrets.toml.**\n\n"
        "Scores will only be saved locally. To enable Google Sheets, add:\n\n"
        "```toml\n"
        'GSHEETS_WEBHOOK_URL = "your-apps-script-url-here"\n'
        "```"
    )

# ── Groq client ───────────────────────────────────────────────────────────────
client = Groq(api_key=GROQ_API_KEY)

def chat_with_llm(messages):
    try:
        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=messages,
            max_tokens=1024,
            temperature=0.7
        )
        return response.choices[0].message.content
    except Exception as e:
        return f"[AI error — please try again: {e}]"

# ── Google Sheets webhook ─────────────────────────────────────────────────────
def save_to_google_sheets(student_name, student_roll, subject, difficulty,
                           score, grade, strengths, improvements,
                           exchanges, project_name, cheated=False):
    if not GSHEETS_WEBHOOK_URL:
        return False
    payload = {
        "timestamp":    datetime.now().strftime("%d-%b-%Y %H:%M"),
        "student_name": student_name,
        "student_roll": student_roll,
        "subject":      subject,
        "difficulty":   difficulty,
        "project_name": project_name,
        "score":        score if score is not None else "Pending",
        "grade":        grade,
        "exchanges":    exchanges,
        "strengths":    strengths,
        "improvements": improvements,
        "cheated":      "Yes" if cheated else "No",
    }
    try:
        r = requests.post(GSHEETS_WEBHOOK_URL, json=payload, timeout=10)
        return r.status_code == 200 and "success" in r.text.lower()
    except Exception as e:
        st.warning(f"Google Sheets post failed: {e}")
        return False

# ── Timer / anti-cheat component ──────────────────────────────────────────────
COMPONENT_DIR = os.path.join(_SCRIPT_DIR, "viva_controller_component")
os.makedirs(COMPONENT_DIR, exist_ok=True)

def render_viva_controller(remaining_seconds):
    html_code = f"""<!DOCTYPE html>
<html>
<head>
<style>
.timer-box {{
    font-family: Arial, sans-serif; background:#ffebee; color:#c62828;
    padding:12px; border-radius:8px; border:1px solid #ffcdd2;
    text-align:center; margin-bottom:15px; font-weight:bold;
}}
.timer-label {{ font-size:10px; text-transform:uppercase; letter-spacing:1px;
    color:#b71c1c; margin-bottom:4px; }}
.timer-value {{ font-size:22px; }}
.fs-box {{
    font-family:Arial,sans-serif; background:#fff8e1;
    border:1px solid #ffe082; border-radius:8px;
    padding:10px; text-align:center; margin-bottom:15px;
}}
.fs-warn {{ font-size:11px; color:#e65100; margin-bottom:6px; }}
.fs-btn {{
    background:#e65100; color:#fff; border:none; border-radius:6px;
    padding:8px 14px; font-weight:bold; font-size:13px;
    cursor:pointer; width:100%;
}}
</style>
<script>
function _post(type, data) {{
    window.parent.postMessage(Object.assign({{isStreamlitMessage:true, type:type}}, data), "*");
}}
function sendValue(value) {{
    _post("streamlit:setComponentValue", {{value:value, dataType:"json"}});
}}
_post("streamlit:componentReady", {{apiVersion:1}});
_post("streamlit:setFrameHeight", {{height:150}});

let timeLeft = {remaining_seconds};
function updateDisplay() {{
    let m = Math.floor(timeLeft/60), s = timeLeft%60;
    if (s < 10) s = "0"+s;
    let el = document.getElementById("timer");
    if (el) el.innerText = m+":"+s;
    if (timeLeft < 60) {{
        let c = document.getElementById("tc");
        if (c) c.style.background="#ffcdd2";
    }}
}}
let iv = setInterval(function() {{
    timeLeft--;
    if (timeLeft <= 0) {{ clearInterval(iv); sendValue("timeout"); }}
    else updateDisplay();
}}, 1000);

let cheatReported = false;
function reportCheat(reason) {{
    if (cheatReported) return;
    cheatReported = true;
    sendValue("cheated");
}}

let gWin, gDoc;
try {{ gWin=window.top; gDoc=window.top.document; void gDoc.visibilityState; }}
catch(e) {{ gWin=window; gDoc=document; }}

document.addEventListener('visibilitychange', function() {{
    if (document.visibilityState==='hidden') reportCheat("tab hidden");
}});
try {{ gDoc.addEventListener('visibilitychange', function() {{
    if (gDoc.visibilityState==='hidden') reportCheat("tab hidden top");
}}); }} catch(e) {{}}

window.addEventListener('blur', function() {{ reportCheat("focus lost"); }});
try {{ gWin.addEventListener('blur', function() {{ reportCheat("focus lost top"); }}); }} catch(e) {{}}

try {{
    gDoc.addEventListener('contextmenu', function(e) {{ e.preventDefault(); }});
    ['copy','cut','paste'].forEach(function(ev) {{
        gDoc.addEventListener(ev, function(e) {{ e.preventDefault(); }});
    }});
}} catch(e) {{}}

function keyGuard(e) {{
    let k=(e.key||"").toLowerCase(), cmd=e.ctrlKey||e.metaKey;
    if (cmd && ['t','n','w','tab'].indexOf(k)!==-1) {{
        e.preventDefault(); reportCheat("new tab/window");
    }}
    if (k==='f12'||(cmd&&e.shiftKey&&['i','j','c'].indexOf(k)!==-1)||(cmd&&k==='u')) {{
        e.preventDefault(); reportCheat("devtools");
    }}
}}
window.addEventListener('keydown', keyGuard, true);
try {{ gWin.addEventListener('keydown', keyGuard, true); }} catch(e) {{}}
try {{ gWin.addEventListener('beforeunload', function(e) {{
    e.preventDefault(); e.returnValue="";
}}); }} catch(e) {{}}

function isFS() {{
    try {{ return !!(gDoc.fullscreenElement||gDoc.webkitFullscreenElement); }}
    catch(e) {{ return !!document.fullscreenElement; }}
}}
function syncFSUI() {{
    let b=document.getElementById("fsbox");
    if (b) b.style.display=isFS()?"none":"block";
}}
function enterFS() {{
    try {{
        let el=gDoc.documentElement;
        let req=el.requestFullscreen||el.webkitRequestFullscreen;
        if (req) {{ req.call(el); try {{ gWin.__vivaFS=true; }} catch(e) {{}} }}
    }} catch(e) {{}}
}}
function fsChange() {{
    let entered=false;
    try {{ entered=!!gWin.__vivaFS; }} catch(e) {{}}
    if (!isFS() && entered) reportCheat("exited fullscreen");
    syncFSUI();
}}
try {{ gDoc.addEventListener('fullscreenchange', fsChange); }} catch(e) {{}}
document.addEventListener('fullscreenchange', fsChange);

window.addEventListener("load", function() {{
    updateDisplay();
    let btn=document.getElementById("fsbtn");
    if (btn) btn.addEventListener("click", enterFS);
    syncFSUI();
}});
</script>
</head>
<body style="margin:0;padding:0;background:transparent;">
<div id="fsbox" class="fs-box" style="display:none;">
    <div class="fs-warn">This viva must run in fullscreen.<br>
    Exiting fullscreen will submit your exam.</div>
    <button id="fsbtn" class="fs-btn">Enter Fullscreen to continue</button>
</div>
<div id="tc" class="timer-box">
    <div class="timer-label">Time Remaining</div>
    <div id="timer" class="timer-value">--:--</div>
</div>
</body>
</html>"""
    with open(os.path.join(COMPONENT_DIR, "index.html"), "w", encoding="utf-8") as f:
        f.write(html_code)
    focus_monitor = components.declare_component("viva_controller", path=COMPONENT_DIR)
    return focus_monitor(key="viva_control_instance")

# ── Text extraction ───────────────────────────────────────────────────────────
def extract_text(uploaded_file):
    text = ""
    try:
        if uploaded_file.name.lower().endswith(".pdf"):
            doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
            for page in doc:
                text += page.get_text()
        elif uploaded_file.name.lower().endswith(".docx"):
            doc = Document(io.BytesIO(uploaded_file.read()))
            for para in doc.paragraphs:
                if para.text.strip():
                    text += para.text + "\n"
    except Exception as e:
        st.error(f"Could not read file: {e}")
    return text.strip()

def truncate(text, max_chars=6000):
    if len(text) > max_chars:
        return text[:max_chars] + "\n\n[... project continues ...]"
    return text

# ── Score / feedback parsers ──────────────────────────────────────────────────
def parse_score(text):
    if not text:
        return None
    cleaned = re.sub(r'[\*_`]', '', text)
    patterns = [
        r'score\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*/\s*10',
        r'(?:final|overall)?\s*score\s*[:\-]?\s*(\d+(?:\.\d+)?)\b',
        r'\b(\d+(?:\.\d+)?)\s*/\s*10\b',
        r'\b(\d+(?:\.\d+)?)\s+out\s+of\s+10\b',
        r'(?:give|rate|award)\s+(?:this|you|the\s+student)?\s*(?:an?\s+)?(\d+(?:\.\d+)?)',
    ]
    for p in patterns:
        m = re.search(p, cleaned, re.IGNORECASE)
        if m:
            try:
                v = float(m.group(1))
                if 0 <= v <= 10:
                    return v
            except ValueError:
                continue
    return None

def parse_feedback(text):
    strengths, improvements = "", ""
    s = re.search(
        r'strength[s]?[:\-\*\s]+(.*?)(?=area|improvement|weakness|recommend|$)',
        text, re.IGNORECASE | re.DOTALL)
    if s:
        strengths = s.group(1).strip()[:300]
    i = re.search(
        r'(?:area[s]?\s*(?:for|needing)|improvement[s]?|weakness)[:\-\*\s]+(.*?)(?=recommend|suggest|score|$)',
        text, re.IGNORECASE | re.DOTALL)
    if i:
        improvements = i.group(1).strip()[:300]
    return strengths, improvements

def get_grade(score):
    if score is None:  return "N/A"
    if score >= 9:     return "O (Outstanding)"
    if score >= 8:     return "A+ (Excellent)"
    if score >= 7:     return "A (Very Good)"
    if score >= 6:     return "B (Good)"
    if score >= 5:     return "C (Average)"
    return                    "F (Fail)"

# ── Excel scoresheet ──────────────────────────────────────────────────────────
def save_to_excel(student_name, student_roll, subject, difficulty, score,
                  grade, strengths, improvements, exchanges,
                  project_name, cheated=False):

    thin   = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    hfont  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hfill  = PatternFill('solid', start_color='1F4E79')

    sfill = (PatternFill('solid', start_color='D9D9D9') if score is None else
             PatternFill('solid', start_color='C6EFCE') if score >= 8 else
             PatternFill('solid', start_color='FFEB9C') if score >= 6 else
             PatternFill('solid', start_color='FFC7CE'))

    if os.path.exists(SCORESHEET_FILE):
        try:
            wb = load_workbook(SCORESHEET_FILE)
            ws = wb.active
            next_row = ws.max_row + 1
        except Exception:
            wb = Workbook(); ws = wb.active; next_row = 2
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Viva Scores"
        headers = ["S.No","Date","Student Name","Roll No","Subject",
                   "Difficulty","Project File","Score (/10)","Grade",
                   "Exchanges","Cheating Flag","Strengths","Areas for Improvement"]
        widths  = [6,18,22,14,22,12,26,12,20,10,14,38,38]
        for col,(h,w) in enumerate(zip(headers,widths),1):
            c = ws.cell(row=1,column=col,value=h)
            c.font=hfont; c.fill=hfill
            c.alignment=center; c.border=border
            ws.column_dimensions[get_column_letter(col)].width=w
        ws.row_dimensions[1].height=30
        ws.freeze_panes="A2"
        next_row=2

    row_data = [
        next_row-1,
        datetime.now().strftime("%d-%b-%Y %H:%M"),
        student_name  or "Unknown",
        student_roll  or "—",
        subject,
        difficulty,
        project_name,
        score if score is not None else "Pending",
        grade,
        exchanges,
        "⚠️ Yes" if cheated else "✅ No",
        strengths     or "—",
        improvements  or "—",
    ]

    for col,value in enumerate(row_data,1):
        c = ws.cell(row=next_row,column=col,value=value)
        c.font=Font(name='Arial',size=10)
        c.border=border
        c.alignment=center if col in [1,8,9,10] else left
        if col==8:
            c.fill=sfill
            c.font=Font(name='Arial',size=10,bold=True)
        elif col==11:
            c.fill=(PatternFill('solid',start_color='FFC7CE') if cheated
                    else PatternFill('solid',start_color='C6EFCE'))
        elif next_row%2==0:
            c.fill=PatternFill('solid',start_color='EBF3FB')
    ws.row_dimensions[next_row].height=40

    summary_row = ws.max_row+2
    stats = [
        ("Total Students", f'=COUNTA(C2:C{next_row})'),
        ("Avg Score",      f'=IFERROR(AVERAGEIF(H2:H{next_row},"<>Pending",H2:H{next_row}),"-")'),
        ("Highest",        f'=IFERROR(MAX(H2:H{next_row}),"-")'),
        ("Lowest",         f'=IFERROR(MINIFS(H2:H{next_row},H2:H{next_row},"<>Pending"),"-")'),
    ]
    for i,(label,formula) in enumerate(stats,1):
        lc=ws.cell(row=summary_row,column=i*2-1,value=label)
        lc.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
        lc.fill=PatternFill('solid',start_color='2E75B6')
        lc.alignment=center
        vc=ws.cell(row=summary_row,column=i*2,value=formula)
        vc.font=Font(name='Arial',bold=True,size=10)
        vc.alignment=center

    try:
        wb.save(SCORESHEET_FILE)
        return SCORESHEET_FILE
    except PermissionError:
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = os.path.join(_SCRIPT_DIR, f"GLIM_Viva_{ts}.xlsx")
        wb.save(alt)
        return alt

# ── Session state ─────────────────────────────────────────────────────────────
DEFAULTS = {
    "messages":              [],
    "viva_started":          False,
    "exchange_count":        0,
    "project_text":          "",
    "viva_ended":            False,
    "score_saved":           False,
    "final_score":           None,
    "final_grade":           "",
    "project_name":          "",
    "closing_message":       "",
    "saved_path":            "",
    "processed_audio_ids":   set(),
    "last_spoken_index":     -1,
    "cheated_detected":      False,
    "timeout_triggered":     False,
    "viva_start_time":       None,
}
for k,v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Viva Settings")

    # Subject and difficulty are locked by the faculty configuration at top
    st.text_input("Subject (Locked)",    value=VIVA_SUBJECT,     disabled=True)
    st.text_input("Difficulty (Locked)", value=VIVA_DIFFICULTY,  disabled=True)

    st.divider()
    st.subheader("👤 Student Details")
    student_name = st.text_input("Full Name",    placeholder="e.g. Priya Sharma")
    student_roll = st.text_input("Roll Number",  placeholder="e.g. PGDM2024-042")

    st.divider()
    st.subheader("🔊 Voice Settings")
    voice_output = st.toggle("🔊 Examiner speaks aloud",      value=True)
    voice_input  = st.toggle("🎙️ Use microphone for answers", value=True)
    if voice_output or voice_input:
        st.success("🎤 Oral viva mode is ON")
    else:
        st.info("📝 Text-only mode")

    st.divider()
    st.subheader("📄 Upload Project")
    uploaded_file = st.file_uploader(
        "Upload project report (PDF or DOCX)",
        type=["pdf","docx"],
        help="Questions will be based on this file"
    )

    ready     = bool(student_name and student_roll and uploaded_file)
    start_btn = st.button(
        "🎓 Start Viva", type="primary",
        disabled=not ready or st.session_state.viva_started
    )
    if not student_name:     st.warning("Enter your full name.")
    elif not student_roll:   st.warning("Enter your roll number.")
    elif not uploaded_file:  st.warning("Upload your project file.")

    if st.session_state.viva_started and not st.session_state.viva_ended:
        st.metric("Questions answered", st.session_state.exchange_count)

# ── Timer / anti-cheat (sidebar injection during active viva) ─────────────────
if (st.session_state.viva_started and not st.session_state.viva_ended
        and VIVA_TIME_LIMIT_MINUTES > 0):
    elapsed   = int(time.time() - (st.session_state.viva_start_time or time.time()))
    remaining = max(0, VIVA_TIME_LIMIT_MINUTES * 60 - elapsed)
    with st.sidebar:
        st.divider()
        signal = render_viva_controller(remaining)
        if signal == "cheated":
            st.session_state.cheated_detected  = True
            st.session_state.viva_ended        = True
            st.rerun()
        elif signal == "timeout" or remaining <= 0:
            st.session_state.timeout_triggered = True
            st.session_state.viva_ended        = True
            st.rerun()

# ── System prompt ─────────────────────────────────────────────────────────────
def build_system_prompt(name, roll, subj, diff, proj_text):
    return f"""You are a strict but fair oral viva examiner for a PGDM {subj} course
at Great Lakes Institute of Management (GLIM), Gurgaon.

Student : {name}  (Roll No: {roll})
Difficulty : {diff}

━━━ PROJECT SUBMITTED ━━━
{truncate(proj_text)}
━━━ END OF PROJECT ━━━

EXAMINATION STRATEGY
════════════════════
PHASE 1 — CONCEPT AUTHENTICATION (Questions 1-3)
  • Identify key concepts, models, and frameworks named in the project above.
  • Ask the student to define each in their own words.
  • Ask how and why that concept was applied in THEIR project.
  • Require confident explanations of at least 2 key concepts before moving on.
  • If the student cannot explain concepts written in their own report,
    note [AUTHENTICITY CONCERN] and weight the final score accordingly.

PHASE 2 — PROJECT DEEP DIVE (Questions 4-10)
  • Probe methodology, data collection, findings, and recommendations.
  • Challenge weak arguments, assumptions, or inconsistencies.
  • Ask at least one subject-knowledge question linked to their project topic.
  • Ask one critical question challenging a gap or weakness you spotted.
  • Final question: "What is the single most important learning from this project?"

RULES
═════
- Ask exactly ONE question at a time (1-2 sentences max — this is spoken).
- Do not repeat questions.
- After question 5 give a one-line mid-viva performance remark.
- After {VIVA_TOTAL_QUESTIONS} questions end with EXACTLY:

VIVA COMPLETE
Score: X/10
Strengths: [2-3 specific points]
Areas for Improvement: [2-3 specific points]
Recommended Topics: [list]

Begin now — greet {name} (Roll No: {roll}) warmly and ask your first
Phase 1 question about a key concept in their project."""

# ── Start viva ────────────────────────────────────────────────────────────────
if start_btn and ready and not st.session_state.viva_started:
    st.session_state.project_name   = uploaded_file.name
    st.session_state.viva_start_time = time.time()
    with st.spinner("📖 Reading project..."):
        project_text = extract_text(uploaded_file)
        st.session_state.project_text = project_text
    if not project_text or len(project_text) < 100:
        st.error("Could not extract text. Check the file is not a scanned image.")
        st.stop()
    prompt = build_system_prompt(
        student_name, student_roll,
        VIVA_SUBJECT, VIVA_DIFFICULTY, project_text)
    st.session_state.messages = [{"role":"system","content":prompt}]
    with st.spinner("Examiner is preparing..."):
        reply = chat_with_llm(st.session_state.messages)
    st.session_state.messages.append({"role":"assistant","content":reply})
    st.session_state.viva_started = True
    st.rerun()

# ── Project preview ───────────────────────────────────────────────────────────
if st.session_state.project_text:
    with st.sidebar:
        st.divider()
        st.caption(f"📋 {st.session_state.project_name}")
        st.caption(f"{len(st.session_state.project_text):,} characters loaded")
        with st.expander("Preview"):
            st.write(st.session_state.project_text[:400]+"...")

# ── Chat display ──────────────────────────────────────────────────────────────
for msg in st.session_state.messages:
    if msg["role"]=="system":
        continue
    with st.chat_message("assistant" if msg["role"]=="assistant" else "user"):
        st.markdown(msg["content"])

# ── Final summary request ─────────────────────────────────────────────────────
def request_final_summary():
    closing = (
        "The viva is now complete. Produce the final assessment immediately "
        "in EXACTLY this format with no extra text:\n\n"
        "VIVA COMPLETE\nScore: X/10\n"
        "Strengths: [2-3 points]\n"
        "Areas for Improvement: [2-3 points]\n"
        "Recommended Topics: [list]"
    )
    st.session_state.messages.append({"role":"user","content":closing})
    summary = chat_with_llm(st.session_state.messages)
    st.session_state.messages.append({"role":"assistant","content":summary})
    return summary

# ── Handle student answer ─────────────────────────────────────────────────────
def handle_response(user_text, via_voice=False):
    st.session_state.messages.append({"role":"user","content":user_text})
    st.session_state.exchange_count += 1

    with st.chat_message("user"):
        st.markdown(f"🎙️ *{user_text}*" if via_voice else user_text)

    with st.chat_message("assistant"):
        with st.spinner("Examiner is responding..."):
            reply = chat_with_llm(st.session_state.messages)
        st.session_state.messages.append({"role":"assistant","content":reply})
        st.markdown(reply)

    st.sidebar.metric("Exchanges", st.session_state.exchange_count)

    should_end = ("VIVA COMPLETE" in reply or
                  st.session_state.exchange_count >= VIVA_TOTAL_QUESTIONS)
    if should_end:
        final_text = reply
        if "VIVA COMPLETE" not in reply:
            with st.chat_message("assistant"):
                with st.spinner("📋 Preparing final assessment..."):
                    final_text = request_final_summary()
                st.markdown(final_text)
        st.session_state.viva_ended      = True
        st.session_state.final_score     = parse_score(final_text)
        st.session_state.final_grade     = get_grade(st.session_state.final_score)
        st.session_state.closing_message = final_text

# ── Input area ────────────────────────────────────────────────────────────────
if st.session_state.viva_started and not st.session_state.viva_ended:
    if voice_input or voice_output:
        bits = []
        if voice_output: bits.append("🔊 examiner speaks")
        if voice_input:  bits.append("🎙️ browser mic enabled")
        st.caption("Voice mode: " + " · ".join(bits))

    if voice_input:
        audio_file = st.audio_input("🎙️ Record your answer", key="viva_audio")
        if audio_file is not None:
            fid = getattr(audio_file,"id",None) or hash(audio_file.getvalue())
            if fid not in st.session_state.processed_audio_ids:
                st.session_state.processed_audio_ids.add(fid)
                with st.spinner("🎙️ Transcribing..."):
                    try:
                        t = client.audio.transcriptions.create(
                            file=(audio_file.name, audio_file.getvalue()),
                            model="whisper-large-v3",
                        )
                        spoken = t.text.strip()
                        if spoken:
                            handle_response(spoken, via_voice=True)
                        else:
                            st.warning("No speech detected. Please try again.")
                    except Exception as e:
                        st.error(f"Transcription failed: {e}")

    typed = st.chat_input("Or type your answer here...")
    if typed:
        handle_response(typed)

# ── Handle cheat / timeout before save ───────────────────────────────────────
if st.session_state.viva_ended and not st.session_state.score_saved:
    if st.session_state.cheated_detected:
        st.session_state.final_score  = 0.0
        st.session_state.final_grade  = "F (Academic Dishonesty)"
        st.session_state.closing_message = "TERMINATED: Tab switching detected."
    elif st.session_state.timeout_triggered:
        with st.spinner("⏳ Time's up — generating assessment..."):
            final = request_final_summary()
        st.session_state.final_score     = parse_score(final)
        st.session_state.final_grade     = get_grade(st.session_state.final_score)
        st.session_state.closing_message = final

# ── Viva ended screen ─────────────────────────────────────────────────────────
if st.session_state.viva_ended:
    st.divider()
    if st.session_state.cheated_detected:
        st.error("🚨 VIVA TERMINATED: Tab switching detected. Incident logged.")
    elif st.session_state.timeout_triggered:
        st.warning("⏱️ Time limit reached. Viva auto-submitted.")
    else:
        st.subheader("🏁 Viva Complete")

    # ── Auto-save scores ──────────────────────────────────────────────────────
    if not st.session_state.score_saved:
        closing = st.session_state.closing_message or (
            st.session_state.messages[-1]["content"]
            if st.session_state.messages else "")
        strengths, improvements = parse_feedback(closing)

        if st.session_state.cheated_detected:
            strengths    = "N/A — terminated for academic dishonesty"
            improvements = "Switched tabs or minimized browser during exam"

        # Local Excel
        try:
            fp = save_to_excel(
                student_name=student_name,
                student_roll=student_roll,
                subject=VIVA_SUBJECT,
                difficulty=VIVA_DIFFICULTY,
                score=st.session_state.final_score,
                grade=st.session_state.final_grade,
                strengths=strengths,
                improvements=improvements,
                exchanges=st.session_state.exchange_count,
                project_name=st.session_state.project_name,
                cheated=st.session_state.cheated_detected,
            )
            st.session_state.saved_path = fp
            st.success(f"✅ Score saved locally: `{fp}`")
        except Exception as e:
            st.error(f"Local save failed: {e}")

        # Google Sheets
        gs_ok = save_to_google_sheets(
            student_name=student_name,
            student_roll=student_roll,
            subject=VIVA_SUBJECT,
            difficulty=VIVA_DIFFICULTY,
            score=st.session_state.final_score,
            grade=st.session_state.final_grade,
            strengths=strengths,
            improvements=improvements,
            exchanges=st.session_state.exchange_count,
            project_name=st.session_state.project_name,
            cheated=st.session_state.cheated_detected,
        )
        if gs_ok:
            st.success("📊 Score posted to Google Sheets!")
        elif GSHEETS_WEBHOOK_URL:
            st.warning("⚠️ Google Sheets post failed — check your webhook URL.")

        st.session_state.score_saved = True

    # ── Score display ─────────────────────────────────────────────────────────
    c1, c2 = st.columns(2)
    with c1:
        st.metric("Final Score",
                  f"{st.session_state.final_score}/10"
                  if st.session_state.final_score is not None else "—")
    with c2:
        st.metric("Grade", st.session_state.final_grade)

    # ── Manual score override ─────────────────────────────────────────────────
    if st.session_state.final_score is None:
        st.warning("No numeric score detected. Enter manually:")
        manual = st.number_input("Manual score (0-10)",
                                 min_value=0.0, max_value=10.0,
                                 step=0.5, value=7.0)
        if st.button("Apply manual score and re-save"):
            st.session_state.final_score = float(manual)
            st.session_state.final_grade = get_grade(st.session_state.final_score)
            closing  = st.session_state.closing_message or ""
            s, i     = parse_feedback(closing)
            fp = save_to_excel(
                student_name=student_name, student_roll=student_roll,
                subject=VIVA_SUBJECT, difficulty=VIVA_DIFFICULTY,
                score=st.session_state.final_score,
                grade=st.session_state.final_grade,
                strengths=s, improvements=i,
                exchanges=st.session_state.exchange_count,
                project_name=st.session_state.project_name,
                cheated=st.session_state.cheated_detected,
            )
            save_to_google_sheets(
                student_name=student_name, student_roll=student_roll,
                subject=VIVA_SUBJECT, difficulty=VIVA_DIFFICULTY,
                score=st.session_state.final_score,
                grade=st.session_state.final_grade,
                strengths=s, improvements=i,
                exchanges=st.session_state.exchange_count,
                project_name=st.session_state.project_name,
                cheated=st.session_state.cheated_detected,
            )
            st.session_state.saved_path = fp
            st.success(f"✅ Updated and saved to `{fp}`")
            st.rerun()

    # ── Download backup ───────────────────────────────────────────────────────
    if st.session_state.saved_path and os.path.exists(st.session_state.saved_path):
        with open(st.session_state.saved_path,"rb") as f:
            st.download_button(
                label="⬇️ Download Scoresheet (backup)",
                data=f,
                file_name="GLIM_Viva_Scores.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

elif not st.session_state.viva_started:
    st.info("👈 Enter your details, upload your project, and click Start Viva.")

# ── Browser TTS ───────────────────────────────────────────────────────────────
if st.session_state.viva_started and voice_output:
    asst_msgs = [m for m in st.session_state.messages if m["role"]=="assistant"]
    if asst_msgs:
        idx = len(asst_msgs)-1
        if st.session_state.last_spoken_index != idx:
            st.session_state.last_spoken_index = idx
            tts_text = asst_msgs[-1]["content"]
            if st.session_state.viva_ended or "VIVA COMPLETE" in tts_text:
                tts_text = ("The viva is now complete. "
                            "Your score and feedback have been recorded. Thank you.")
            escaped = (tts_text
                       .replace('\\','\\\\')
                       .replace('"','\\"')
                       .replace("'","\\'")
                       .replace('\n',' '))
            components.html(f"""<script>
if (window.speechSynthesis) {{
    window.speechSynthesis.cancel();
    setTimeout(function() {{
        var u = new SpeechSynthesisUtterance("{escaped}");
        u.rate = 1.05;
        window.speechSynthesis.speak(u);
    }}, 50);
}}
</script>""", height=0)
